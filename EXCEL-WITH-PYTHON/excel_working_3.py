from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Sample data with student grades
data = {
    "Joe": {"math": 65, "science": 78, "english": 98, "gym": 89},
    "Bill": {"math": 55, "science": 72, "english": 87, "gym": 95},
    "Tim": {"math": 100, "science": 45, "english": 75, "gym": 92},
    "Sally": {"math": 30, "science": 25, "english": 45, "gym": 100},
    "Jane": {"math": 100, "science": 100, "english": 100, "gym": 60},
}

# Create a new workbook and set the active worksheet
wb = Workbook()
ws = wb.active
ws.title = "Grades"

# Write headers to the worksheet
headings = ["Name"] + list(data["Joe"].keys())
ws.append(headings)

# Write student grades to the worksheet
for person in data:
    grades = list(data[person].values())  # Get the list of grades for the current student
    ws.append([person] + grades)  # Append a new row to the worksheet with student name and grades

# Calculate and write averages for each subject
for col in range(2, len(data["Joe"]) + 2):
    char = get_column_letter(col)  # Get the column letter (e.g., 'B', 'C') based on the column index
    ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"  # Calculate and write the average formula

# Apply formatting to header cells (bold and blue font)
for col in range(1, 6):
    header_cell = ws[get_column_letter(col) + "1"]  # Get the header cell in the first row for the current column
    header_cell.font = Font(bold=True, color="0099CCFF")  # Apply formatting (bold and blue font) to the header cell


# Save the workbook to a file
wb.save("EXCEL-WITH-PYTHON/NewGrades.xlsx")
