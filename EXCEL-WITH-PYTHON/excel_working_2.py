from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# wb = Workbook()
# ws = wb.active
# ws.title = "Data"
# ws.append(['Tim', 'Is', 'Great', '!'])
# ws.append(['Tim', 'Is', 'Great', '!'])
# ws.append(['Tim', 'Is', 'Great', '!'])
# ws.append(['Tim', 'Is', 'Great', '!'])
# ws.append(["end"])
# wb.save('EXCEL-WITH-PYTHON/tim.xlsx')

# Load the existing workbook
wb = load_workbook('EXCEL-WITH-PYTHON/tim.xlsx')
ws = wb.active

# Iterate over rows (1 to 10)
for row in range(1, 11):
    # Iterate over columns (1 to 4)
    for col in range(1, 5):
        # Get the column letter for the current column index
        char = get_column_letter(col)
        
        # Form a cell reference (e.g., 'A1') and assign a value
        # The value is a combination of the column letter and row number
        ws[char + str(row)] = char + str(row)

# Save the workbook with the filled cells to 'tim.xlsx'
wb.save('EXCEL-WITH-PYTHON/tim.xlsx')

# ws.merge_cells("A1:D2")  # Merge cells A1 to D2

# ws.insert_rows(7)  # Insert a row at index 7
# ws.insert_rows(7)  # Insert another row at index 7

# ws.delete_rows(7)  # Delete the row at index 7

# ws.insert_cols(2)  # Insert a column at index 2
# ws.delete_cols(2)  # Delete the column at index 2

# ws.move_range("C1:D11", rows=2, cols=2)  # Move the range C1:D11 by 2 rows and 2 columns

# Save the changes to the workbook
wb.save('EXCEL-WITH-PYTHON/tim.xlsx')