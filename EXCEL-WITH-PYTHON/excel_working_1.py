from openpyxl import Workbook, load_workbook

# Load the existing workbook
wb = load_workbook('EXCEL-WITH-PYTHON/Grades.xlsx')

# Get the active sheet
ws = wb.active

# Print the entire worksheet
# print(ws)

# Print the value in cell A1
# print(ws['A1'].value)

# Update the value in cell A2
# ws['A2'].value = "Test"

# Save the changes to the workbook
# wb.save('EXCEL-WITH-PYTHON/Grades.xlsx')

# Print the sheet names in the workbook
# print(wb.sheetnames)

# Access a specific sheet by name
# sheet2 = wb['Sheet2']
# print(sheet2)

# Create a new sheet named "Test"
# wb.create_sheet("Test")

# Print the updated list of sheet names
# print(wb.sheetnames)
